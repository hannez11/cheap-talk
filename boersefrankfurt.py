from requests_html import HTMLSession
import openpyxl
import time
import os
import urllib.request


class Parser:
    def __init__(self, inputfile, outputfile):
        self.inputfile = inputfile
        self.outputfile = outputfile


    def load_xlsx(self):
        print(f"File {self.inputfile} opened")
        self.wb = openpyxl.load_workbook(self.inputfile, data_only=True) #parses the actual values and not the formulas
        self.wb.create_sheet("Downloadlinks")
        self.ws = self.wb["Downloadlinks"]

        for i in range(5): #create new headers for current worksheet
            self.ws.cell(column = i+1, row = 1).value = ["Firmnumber","ISIN","Date","Titel","Link"][i]


    def save_xlsx(self):
        self.wb.save(self.outputfile)
        print(f"File {self.outputfile} saved\n")


    def companies(self): #get pdf download links. download urls apparently change after a few days?!
        self.firmnumber = 1
        self.ws = self.wb["file"] #set active sheet where company links are stored

        start, end = 2, self.wb["file"].max_row #from 2 - 377
        for col in self.ws.iter_cols(min_row=start, max_row=end, min_col=3, max_col=3): #iterate through isin-link column (3) - only adjust max_row
            for cell in col:
                self.url = cell.value #eg https://www.boerse-frankfurt.de/aktie/recent-reports/DE0005545503
                print(f"Current URL of firm #{self.firmnumber}: {self.url}")
                self.ws = self.wb["Downloadlinks"] #jump to new worksheet to save new info
                self.parse_downloadlinks() #runs method to parse the url for a specific ISIN
                self.ws = self.wb["file"] #jump to old worksheet to grab isin-link
                self.save_xlsx() #save file after each firm
            print("All done")
                
    def parse_downloadlinks(self): 
        time.sleep(2)
        session = HTMLSession() #has to be recalled for each page, otherwise closed session error
        r = session.get(self.url)
        r.html.render(timeout=0, sleep=1) #sleep needed for table (js/ajax) elements to load

        table = r.html.find("tr") #find all table rows (last table row has no content)

        for rows in range(1,len(table)): #parse date, doctitel and url for all rows
            try:
                current_row = self.ws.max_row + 1 #use the next empty row to paste info in       
                
                self.ws.cell(column = 1, row = current_row).value = self.firmnumber
                
                date = r.html.find(f'tr.ng-star-inserted:nth-child({rows}) > td:nth-child(1)') #creates a list with 1 entry
                self.ws.cell(column = 3, row = current_row).value = date[0].text 
                #print(date[0].text)

                doctitel = r.html.find(f'tr.ng-star-inserted:nth-child({rows}) > td:nth-child(2)')
                self.ws.cell(column = 4, row = current_row).value = doctitel[0].text
                #print(doctitel[0].text)

                url = r.html.find(f'tr.ng-star-inserted:nth-child({rows}) > td:nth-child(3) > a:nth-child(1)')
                self.ws.cell(column = 5, row = current_row).value = list(url[0].absolute_links)[0]
                #print(list(url[0].absolute_links)[0]) #absolute_links creates a set -> transform to list to access link or [print(rows) for rows in about[0].absolute_links]

                self.ws.cell(column = 2, row = current_row).value = self.ws.cell(column = 5, row = current_row).value.split("=")[1][:-3] #isin column
            except:
                print("Error while going through the table rows")

        session.close() #close current chromium session
        print(f"Firm {self.firmnumber} done")
        self.firmnumber += 1

    
    def downloader(self): #download pdf files
        folder = os.path.join(os.getcwd(), "Desktop", "Jahresabschluss") #C:\Users\hannez\Desktop\Jahresabschluss

        for i in range(1997, 2020): #create 23 folders
            if not os.path.exists(os.path.join(folder, str(i))):
                os.makedirs(os.path.join(folder, str(i)))

        self.wb = openpyxl.load_workbook(self.outputfile, data_only=True)
        self.ws = self.wb["Jahresabschluss"] #manually create new ws Jahresabschluss and copy the values from Downloadlinks in excel: Firmnumber ISIN Date Month PublicationYear BusinessYear Titel Link
        self.ws.cell(column = 9, row = 1).value = "Status" #create new column with header "status"
        
        start, end = 2, self.wb["Jahresabschluss"].max_row #2, self.wb["Jahresabschluss"].max_row
        for col in self.ws.iter_cols(min_row=start, max_row=end, min_col=8, max_col=8): #iterate through link column (8)
            for index, cell in enumerate(col):
                self.link = cell.value #set downloadlurl variable
                year = str(self.ws.cell(column = 6, row = start+index).value) #year variable for naming the pdf file and choosing the folder
                isin = str(self.ws.cell(column = 2, row = start+index).value)
                urllib.request.urlretrieve(self.link, os.path.join(folder, year, f"{isin}_{year}.pdf"))
                print(f"File {isin}_{year}.pdf downloaded\n{index+1} of {end-start+1} downloaded")
                self.ws.cell(column = 9, row = start+index).value = "downloaded" #set status of current row to downloaded

            self.save_xlsx()
            print("All done")


doc1 = Parser(inputfile = "C://Users//hannez//Desktop//CDAX.xlsx", outputfile = "C://Users//hannez//Desktop//CDAX2020.xlsx")

#1 get links
# doc1.load_xlsx()
# doc1.companies()
# doc1.save_xlsx()

#2 download pdfs
doc1.downloader()